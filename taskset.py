import numpy as np
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
import os

epsilon = 1e-9
SIMULATION_TIME = 50
NUM_CORES = 2
ALPHA = 0.8
BETA = 0.2
GAMMA = 0
CONTEXT_SWITCH_TIME = 1  # 1 unit for each add/remove operation
current_time = 0
preemptions = 0
deadline_misses = 0
history = []
prange = 0


class Task:
    min_laxity = float('inf')
    max_laxity = float('-inf')

    def __init__(self, id, arrival_time, burst_time, deadline, priority):
        self.relaxation = None
        self.id = id
        self.arrival_time = arrival_time
        self.burst_time = burst_time
        self.deadline = deadline
        self.priority = priority
        self.remaining_time = burst_time
        self.laxity = deadline - burst_time
        self.dropped = False
        self.completion_time = None

    def update_laxity(self):
        self.laxity = self.deadline - self.remaining_time - current_time
        return self.laxity

    def normalized_laxity(self):
        return (self.update_laxity() - Task.min_laxity) * prange / (Task.max_laxity - Task.min_laxity + epsilon)

    def update_relaxation(self):
        return ALPHA * self.normalized_laxity() + BETA * self.priority + GAMMA * self.deadline

    def __lt__(self, other):
        return self.deadline < other.deadline

    def __repr__(self):
        return f'Task: {self.id}, DL: {self.deadline}, PR: {self.priority}, remaining: {self.remaining_time},'

tasks = [
    # Task sets with size = 5
    [
        Task(1, 0, 2, 8, 1),
        Task(2, 1, 3, 10, 2),
        Task(3, 2, 1, 6, 3),
        Task(4, 3, 2, 9, 4),
        Task(5, 4, 4, 7, 1)
    ],
    [
        Task(1, 0, 3, 10, 2),
        Task(2, 1, 4, 8, 4),
        Task(3, 2, 1, 10, 1),
        Task(4, 3, 2, 8, 3),
        Task(5, 4, 3, 11, 2)
    ],
    [
        Task(1, 0, 4, 7, 3),
        Task(2, 1, 3, 11, 2),
        Task(3, 2, 5, 7, 4),
        Task(4, 3, 5, 15, 1),
        Task(5, 4, 3, 10, 2)
    ],

    # Task sets with size = 6
    [
        Task(1, 0, 2, 8, 1),
        Task(2, 1, 3, 10, 2),
        Task(3, 2, 1, 6, 3),
        Task(4, 3, 2, 9, 4),
        Task(5, 4, 4, 12, 1),
        Task(6, 5, 3, 11, 2)
    ],
    [
        Task(1, 0, 3, 10, 2),
        Task(2, 1, 4, 13, 4),
        Task(3, 2, 1, 6, 1),
        Task(4, 3, 2, 8, 3),
        Task(5, 4, 3, 11, 2),
        Task(6, 5, 2, 9, 3)
    ],
    [
        Task(1, 0, 4, 12, 3),
        Task(2, 1, 3, 11, 2),
        Task(3, 2, 4, 7, 4),
        Task(4, 3, 2, 9, 1),
        Task(5, 4, 7, 10, 2),
        Task(6, 5, 4, 14, 3)
    ],

    # Task sets with size = 7
    [
        Task(1, 0, 2, 8, 1),
        Task(2, 1, 3, 10, 2),
        Task(3, 2, 1, 6, 3),
        Task(4, 3, 2, 9, 4),
        Task(5, 4, 4, 12, 1),
        Task(6, 5, 8, 11, 2),
        Task(7, 6, 2, 9, 3)
    ],
    [
        Task(1, 0, 3, 10, 2),
        Task(2, 1, 4, 13, 4),
        Task(3, 2, 1, 6, 1),
        Task(4, 3, 2, 8, 3),
        Task(5, 4, 3, 11, 2),
        Task(6, 5, 2, 9, 3),
        Task(7, 6, 1, 7, 4)
    ],
    [ #this is the sample
        Task(1, 0, 4, 12, 3),
        Task(2, 1, 3, 11, 2),
        Task(3, 2, 1, 7, 4),
        Task(4, 3, 2, 9, 1),
        Task(5, 4, 3, 10, 2),
        Task(6, 5, 4, 14, 3),
        Task(7, 6, 3, 12, 2)
    ],

    # Task sets with size = 8
    [
        Task(1, 0, 2, 8, 1),
        Task(2, 1, 6, 10, 2),
        Task(3, 2, 3, 6, 3),
        Task(4, 3, 2, 9, 4),
        Task(5, 4, 4, 12, 1),
        Task(6, 5, 8, 11, 2),
        Task(7, 6, 2, 9, 3),
        Task(8, 7, 1, 7, 4)
    ],
    [
        Task(1, 0, 3, 10, 2),
        Task(2, 1, 4, 13, 4),
        Task(3, 2, 1, 6, 1),
        Task(4, 3, 2, 8, 3),
        Task(5, 4, 6, 11, 2),
        Task(6, 5, 2, 9, 3),
        Task(7, 6, 4, 7, 4),
        Task(8, 7, 6, 12, 3)
    ],
    [
        Task(1, 0, 4, 12, 3),
        Task(2, 1, 3, 11, 2),
        Task(3, 2, 1, 7, 4),
        Task(4, 3, 2, 9, 1),
        Task(5, 4, 3, 10, 2),
        Task(6, 5, 4, 14, 3),
        Task(7, 6, 3, 12, 2),
        Task(8, 7, 1, 6, 4)
    ],

    # Task sets with size = 9
    [
        Task(1, 0, 3, 10, 2),
        Task(2, 1, 4, 13, 4),
        Task(3, 2, 1, 6, 3),
        Task(4, 3, 2, 8, 1),
        Task(5, 4, 3, 11, 2),
        Task(6, 5, 6, 7, 4),
        Task(7, 6, 2, 9, 3),
        Task(8, 7, 6, 12, 1),
        Task(9, 8, 9, 15, 2)
    ],
    [
        Task(1, 0, 4, 12, 1),
        Task(2, 1, 2, 9, 2),
        Task(3, 2, 3, 10, 3),
        Task(4, 3, 1, 7, 4),
        Task(5, 4, 3, 11, 1),
        Task(6, 5, 4, 13, 3),
        Task(7, 6, 2, 9, 2),
        Task(8, 7, 1, 6, 4),
        Task(9, 8, 3, 14, 1)
    ],
    [
        Task(1, 0, 2, 9, 3),
        Task(2, 1, 5, 13, 1),
        Task(3, 2, 4, 6, 2),
        Task(4, 3, 7, 10, 4),
        Task(5, 4, 2, 7, 3),
        Task(6, 5, 3, 12, 1),
        Task(7, 6, 3, 10, 2),
        Task(8, 7, 3, 14, 4),
        Task(9, 8, 2, 13, 3)
    ],

    # Task sets with size = 10
    [
        Task(1, 0, 3, 10, 2),
        Task(2, 1, 4, 13, 1),
        Task(3, 2, 1, 6, 4),
        Task(4, 3, 3, 11, 2),
        Task(5, 4, 2, 7, 3),
        Task(6, 5, 4, 12, 1),
        Task(7, 6, 6, 19, 3),
        Task(8, 7, 1, 9, 4),
        Task(9, 8, 9, 14, 2),
        Task(10, 9, 4, 15, 1)
    ],
    [
        Task(1, 0, 2, 9, 3),
        Task(2, 1, 4, 12, 2),
        Task(3, 2, 4, 7, 4),
        Task(4, 3, 3, 10, 1),
        Task(5, 4, 2, 8, 3),
        Task(6, 5, 3, 11, 2),
        Task(7, 6, 6, 12, 1),
        Task(8, 7, 1, 6, 4),
        Task(9, 8, 4, 13, 3),
        Task(10, 9, 5, 14, 4)
    ],
    [
        Task(1, 0, 3, 11, 1),
        Task(2, 1, 2, 9, 2),
        Task(3, 2, 4, 13, 3),
        Task(4, 3, 6, 6, 4),
        Task(5, 4, 3, 10, 2),
        Task(6, 5, 2, 8, 1),
        Task(7, 6, 5, 14, 3),
        Task(8, 7, 3, 7, 2),
        Task(9, 8, 7, 12, 4),
        Task(10, 9, 2, 13, 1)
    ],
    [
        Task(1, 0, 6, 10, 2),
        Task(2, 1, 3, 14, 1),
        Task(3, 2, 2, 12, 3),
        Task(4, 3, 5, 13, 2),
        Task(5, 4, 3, 9, 1),
        Task(6, 5, 6, 15, 4),
        Task(7, 6, 2, 11, 3),
        Task(8, 7, 7, 16, 1),
        Task(9, 10, 2, 10, 4),
        Task(10, 11, 3, 17, 2),
        Task(11, 13, 1, 8, 3)
    ],
    [
        Task(1, 0, 4, 14, 1),
        Task(2, 1, 3, 13, 2),
        Task(3, 2, 2, 12, 3),
        Task(4, 3, 1, 11, 4),
        Task(5, 4, 3, 15, 2),
        Task(6, 5, 4, 16, 3),
        Task(7, 6, 2, 10, 4),
        Task(8, 7, 3, 17, 1),
        Task(9, 8, 2, 12, 2),
        Task(10, 9, 3, 14, 4),
        Task(11, 10, 1, 9, 3)
    ],
    [
        Task(1, 0, 3, 15, 2),
        Task(2, 1, 4, 14, 1),
        Task(3, 2, 1, 10, 3),
        Task(4, 3, 3, 16, 4),
        Task(5, 4, 2, 12, 3),
        Task(6, 5, 9, 17, 2),
        Task(7, 6, 5, 9, 4),
        Task(8, 7, 10, 18, 1),
        Task(9, 8, 6, 11, 2),
        Task(10, 9, 4, 19, 3),
        Task(11, 10, 6, 13, 4)
    ],

    # Task sets with size = 12
    [
        Task(1, 0, 3, 14, 2),
        Task(2, 1, 4, 15, 1),
        Task(3, 2, 2, 12, 3),
        Task(4, 3, 3, 16, 4),
        Task(5, 4, 1, 9, 2),
        Task(6, 5, 4, 11, 3),
        Task(7, 6, 2, 8, 4),
        Task(8, 7, 3, 10, 2),
        Task(9, 8, 2, 18, 1),
        Task(10, 9, 4, 16, 3),
        Task(11, 10, 1, 5, 4),
        Task(12, 11, 3, 20, 1)
    ],
    [
        Task(1, 0, 3, 13, 1),
        Task(2, 1, 2, 10, 4),
        Task(3, 2, 13, 15, 2),
        Task(4, 3, 6, 14, 3),
        Task(5, 4, 4, 17, 1),
        Task(6, 5, 2, 12, 4),
        Task(7, 6, 3, 16, 2),
        Task(8, 7, 5, 15, 3),
        Task(9, 8, 2, 11, 1),
        Task(10, 9, 1, 8, 4),
        Task(11, 10, 4, 19, 2),
        Task(12, 11, 9, 18, 3)
    ],
    [
        Task(1, 0, 4, 12, 3),
        Task(2, 1, 3, 4, 2),
        Task(3, 2, 9, 10, 4),
        Task(4, 3, 12, 13, 1),
        Task(5, 4, 2, 19, 2),
        Task(6, 5, 4, 17, 3),
        Task(7, 6, 2, 16, 4),
        Task(8, 7, 3, 10, 2),
        Task(9, 8, 1, 9, 1),
        Task(10, 9, 2, 9, 3),
        Task(11, 10, 3, 20, 4),
        Task(12, 11, 4, 17, 2)
    ],

    # Task sets with size = 13
    [
        Task(1, 0, 3, 12, 2),
        Task(2, 1, 4, 14, 1),
        Task(3, 2, 2, 11, 3),
        Task(4, 3, 3, 6, 4),
        Task(5, 4, 1, 3, 2),
        Task(6, 5, 4, 18, 1),
        Task(7, 6, 2, 12, 3),
        Task(8, 7, 3, 17, 4),
        Task(9, 8, 3, 9, 2),
        Task(10, 9, 4, 19, 3),
        Task(11, 10, 2, 13, 4),
        Task(12, 11, 3, 20, 1),
        Task(13, 12, 1, 8, 2)
    ],
    [
        Task(1, 0, 2, 10, 3),
        Task(2, 1, 4, 14, 1),
        Task(3, 2, 1, 12, 4),
        Task(4, 3, 3, 16, 2),
        Task(5, 4, 2, 11, 3),
        Task(6, 5, 4, 18, 1),
        Task(7, 6, 3, 15, 2),
        Task(8, 7, 2, 13, 4),
        Task(9, 8, 4, 17, 1),
        Task(10, 9, 1, 8, 3),
        Task(11, 10, 3, 20, 2),
        Task(12, 11, 2, 14, 4),
        Task(13, 12, 1, 9, 3)
    ],
    [
        Task(1, 0, 3, 13, 1),
        Task(2, 1, 2, 11, 2),
        Task(3, 2, 3, 15, 3),
        Task(4, 3, 1, 10, 4),
        Task(5, 4, 4, 18, 2),
        Task(6, 5, 2, 14, 3),
        Task(7, 6, 3, 16, 1),
        Task(8, 7, 1, 9, 4),
        Task(9, 8, 4, 19, 2),
        Task(10, 9, 2, 13, 3),
        Task(11, 10, 1, 8, 4),
        Task(12, 11, 3, 17, 1),
        Task(13, 12, 4, 20, 2)
    ],

    # Task sets with size = 14
    [
        Task(1, 0, 3, 14, 2),
        Task(2, 1, 4, 15, 1),
        Task(3, 2, 2, 12, 3),
        Task(4, 3, 3, 16, 4),
        Task(5, 4, 1, 9, 2),
        Task(6, 5, 4, 17, 3),
        Task(7, 6, 2, 11, 4),
        Task(8, 7, 3, 18, 2),
        Task(9, 8, 2, 13, 1),
        Task(10, 9, 4, 19, 3),
        Task(11, 10, 1, 10, 4),
        Task(12, 11, 3, 20, 1),
        Task(13, 12, 2, 14, 2),
        Task(14, 13, 4, 21, 3)
    ],
    [
        Task(1, 0, 4, 15, 1),
        Task(2, 1, 3, 13, 3),
        Task(3, 2, 2, 12, 2),
        Task(4, 3, 1, 10, 4),
        Task(5, 4, 3, 14, 1),
        Task(6, 5, 4, 17, 2),
        Task(7, 6, 1, 9, 3),
        Task(8, 7, 2, 16, 4),
        Task(9, 8, 3, 18, 1),
        Task(10, 9, 1, 8, 2),
        Task(11, 10, 4, 20, 3),
        Task(12, 11, 2, 13, 4),
        Task(13, 12, 3, 17, 2),
        Task(14, 13, 1, 7, 3)
    ],
    [
        Task(1, 0, 3, 14, 2),
        Task(2, 1, 4, 16, 1),
        Task(3, 2, 2, 12, 3),
        Task(4, 3, 3, 17, 4),
        Task(5, 4, 1, 9, 2),
        Task(6, 5, 4, 18, 3),
        Task(7, 6, 2, 11, 4),
        Task(8, 7, 3, 19, 2),
        Task(9, 8, 1, 8, 1),
        Task(10, 9, 2, 14, 3),
        Task(11, 10, 3, 21, 4),
        Task(12, 11, 4, 22, 1),
        Task(13, 12, 2, 13, 3),
        Task(14, 13, 1, 10, 2)
    ],
    # Task set 1 with size = 15
    [
        Task(1, 0, 3, 12, 1),
        Task(2, 1, 4, 15, 3),
        Task(3, 2, 2, 11, 4),
        Task(4, 3, 1, 10, 2),
        Task(5, 4, 3, 14, 1),
        Task(6, 5, 4, 16, 4),
        Task(7, 6, 2, 13, 2),
        Task(8, 7, 3, 17, 3),
        Task(9, 8, 1, 9, 1),
        Task(10, 9, 4, 18, 2),
        Task(11, 10, 3, 19, 4),
        Task(12, 11, 2, 20, 3),
        Task(13, 12, 4, 21, 1),
        Task(14, 13, 1, 8, 2),
        Task(15, 14, 3, 22, 4)
    ],
    # Task set 2 with size = 15
    [
        Task(1, 0, 2, 10, 3),
        Task(2, 1, 3, 13, 1),
        Task(3, 2, 4, 15, 4),
        Task(4, 3, 2, 11, 2),
        Task(5, 4, 3, 14, 1),
        Task(6, 5, 4, 16, 3),
        Task(7, 6, 1, 9, 2),
        Task(8, 7, 2, 12, 4),
        Task(9, 8, 3, 17, 1),
        Task(10, 9, 4, 18, 2),
        Task(11, 10, 2, 19, 3),
        Task(12, 11, 3, 20, 4),
        Task(13, 12, 1, 8, 2),
        Task(14, 13, 4, 22, 1),
        Task(15, 14, 3, 23, 4)
    ],
    # Task set 3 with size = 15
    [
        Task(1, 0, 4, 16, 2),
        Task(2, 1, 3, 14, 4),
        Task(3, 2, 1, 10, 3),
        Task(4, 3, 2, 11, 1),
        Task(5, 4, 3, 15, 2),
        Task(6, 5, 1, 9, 4),
        Task(7, 6, 4, 17, 3),
        Task(8, 7, 2, 12, 1),
        Task(9, 8, 3, 18, 2),
        Task(10, 9, 1, 8, 4),
        Task(11, 10, 4, 19, 3),
        Task(12, 11, 2, 13, 1),
        Task(13, 12, 3, 20, 4),
        Task(14, 13, 1, 7, 2),
        Task(15, 14, 2, 21, 3)
    ],
    [
        Task(1, 0, 3, 15, 2),
        Task(2, 1, 4, 16, 1),
        Task(3, 2, 2, 14, 3),
        Task(4, 3, 3, 17, 4),
        Task(5, 4, 1, 12, 2),
        Task(6, 5, 4, 18, 3),
        Task(7, 6, 2, 11, 4),
        Task(8, 7, 3, 19, 2),
        Task(9, 8, 1, 10, 1),
        Task(10, 9, 4, 20, 3),
        Task(11, 10, 2, 13, 4),
        Task(12, 11, 3, 21, 1),
        Task(13, 12, 4, 22, 2),
        Task(14, 13, 1, 9, 3),
        Task(15, 14, 3, 23, 1),
        Task(16, 15, 2, 24, 4)
    ],
    [
        Task(1, 0, 4, 16, 1),
        Task(2, 1, 2, 14, 4),
        Task(3, 2, 3, 17, 2),
        Task(4, 3, 1, 10, 3),
        Task(5, 4, 4, 18, 2),
        Task(6, 5, 2, 15, 1),
        Task(7, 6, 3, 19, 4),
        Task(8, 7, 1, 12, 3),
        Task(9, 8, 2, 20, 1),
        Task(10, 9, 4, 21, 2),
        Task(11, 10, 3, 22, 4),
        Task(12, 11, 1, 11, 2),
        Task(13, 12, 3, 23, 1),
        Task(14, 13, 2, 24, 3),
        Task(15, 14, 1, 13, 4),
        Task(16, 15, 4, 25, 2)
    ],
    [
        Task(1, 0, 2, 11, 3),
        Task(2, 1, 4, 15, 2),
        Task(3, 2, 3, 14, 1),
        Task(4, 3, 1, 10, 4),
        Task(5, 4, 2, 13, 2),
        Task(6, 5, 4, 18, 1),
        Task(7, 6, 2, 12, 3),
        Task(8, 7, 3, 16, 4),
        Task(9, 8, 1, 9, 2),
        Task(10, 9, 2, 17, 3),
        Task(11, 10, 4, 19, 1),
        Task(12, 11, 3, 20, 4),
        Task(13, 12, 2, 21, 3),
        Task(14, 13, 1, 8, 2),
        Task(15, 14, 3, 22, 4),
        Task(16, 15, 2, 23, 1)
    ],

    # Task sets with size = 17
    [
        Task(1, 0, 3, 14, 1),
        Task(2, 1, 4, 16, 2),
        Task(3, 2, 2, 13, 4),
        Task(4, 3, 1, 10, 3),
        Task(5, 4, 4, 18, 2),
        Task(6, 5, 3, 17, 1),
        Task(7, 6, 1, 12, 4),
        Task(8, 7, 2, 15, 3),
        Task(9, 8, 4, 20, 1),
        Task(10, 9, 3, 19, 2),
        Task(11, 10, 1, 9, 3),
        Task(12, 11, 2, 21, 4),
        Task(13, 12, 4, 22, 2),
        Task(14, 13, 3, 23, 1),
        Task(15, 14, 1, 8, 4),
        Task(16, 15, 2, 24, 3),
        Task(17, 16, 3, 25, 3)
    ],
    [
        Task(1, 0, 2, 12, 3),
        Task(2, 1, 3, 14, 2),
        Task(3, 2, 1, 10, 4),
        Task(4, 3, 4, 16, 1),
        Task(5, 4, 2, 13, 3),
        Task(6, 5, 4, 18, 2),
        Task(7, 6, 3, 17, 4),
        Task(8, 7, 1, 9, 2),
        Task(9, 8, 3, 15, 1),
        Task(10, 9, 4, 19, 3),
        Task(11, 10, 2, 12, 4),
        Task(12, 11, 3, 20, 2),
        Task(13, 12, 4, 21, 1),
        Task(14, 13, 1, 8, 3),
        Task(15, 14, 2, 22, 4),
        Task(16, 15, 3, 23, 1),
        Task(17, 4, 6, 23, 1),
    ],
    [
        Task(1, 0, 3, 14, 2),
        Task(2, 1, 2, 12, 4),
        Task(3, 2, 3, 15, 1),
        Task(4, 3, 4, 18, 3),
        Task(5, 4, 1, 9, 2),
        Task(6, 5, 4, 16, 1),
        Task(7, 6, 2, 11, 4),
        Task(8, 7, 3, 19, 2),
        Task(9, 8, 1, 10, 3),
        Task(10, 9, 4, 20, 1),
        Task(11, 10, 2, 13, 2),
        Task(12, 11, 3, 21, 4),
        Task(13, 12, 1, 8, 3),
        Task(14, 13, 4, 22, 2),
        Task(15, 14, 2, 23, 4),
        Task(16, 15, 3, 24, 1),
        Task(17, 19, 3, 6, 4),
    ],

    # Task sets with size = 18
    [
        Task(1, 0, 2, 12, 4),
        Task(2, 1, 3, 14, 2),
        Task(3, 2, 1, 10, 3),
        Task(4, 3, 4, 16, 1),
        Task(5, 4, 3, 15, 2),
        Task(6, 5, 2, 13, 4),
        Task(7, 6, 3, 17, 3),
        Task(8, 7, 1, 9, 2),
        Task(9, 8, 4, 18, 1),
        Task(10, 9, 2, 11, 4),
        Task(11, 10, 3, 19, 2),
        Task(12, 11, 1, 8, 3),
        Task(13, 12, 4, 20, 4),
        Task(14, 13, 3, 21, 1),
        Task(15, 14, 2, 23, 2),
        Task(16, 15, 4, 22, 3),
        Task(17, 16, 1, 7, 2),
        Task(18, 17, 3, 24, 4)
    ],
    [
        Task(1, 0, 3, 14, 2),
        Task(2, 1, 4, 16, 3),
        Task(3, 2, 1, 10, 4),
        Task(4, 3, 2, 13, 1),
        Task(5, 4, 3, 15, 4),
        Task(6, 5, 2, 12, 2),
        Task(7, 6, 4, 17, 1),
        Task(8, 7, 1, 9, 3),
        Task(9, 8, 3, 18, 2),
        Task(10, 9, 4, 19, 1),
        Task(11, 10, 2, 11, 3),
        Task(12, 11, 3, 20, 4),
        Task(13, 12, 1, 8, 2),
        Task(14, 13, 4, 22, 1),
        Task(15, 14, 3, 23, 4),
        Task(16, 15, 2, 24, 3),
        Task(17, 16, 1, 7, 4),
        Task(18, 17, 3, 25, 2)
    ],
    [
        Task(1, 0, 4, 16, 1),
        Task(2, 1, 3, 14, 4),
        Task(3, 2, 2, 13, 2),
        Task(4, 3, 1, 10, 3),
        Task(5, 4, 4, 17, 1),
        Task(6, 5, 2, 12, 2),
        Task(7, 6, 3, 18, 3),
        Task(8, 7, 1, 9, 4),
        Task(9, 8, 4, 19, 2),
        Task(10, 9, 3, 15, 1),
        Task(11, 10, 2, 11, 4),
        Task(12, 11, 3, 20, 2),
        Task(13, 12, 1, 8, 3),
        Task(14, 13, 4, 22, 1),
        Task(15, 14, 3, 21, 2),
        Task(16, 15, 1, 7, 4),
        Task(17, 16, 4, 23, 3),
        Task(18, 17, 2, 24, 1)
    ],

    # Task set 1 with size = 19
    [
        Task(1, 0, 2, 11, 1),
        Task(2, 1, 4, 16, 3),
        Task(3, 2, 3, 13, 4),
        Task(4, 3, 2, 12, 2),
        Task(5, 4, 3, 14, 1),
        Task(6, 5, 1, 10, 4),
        Task(7, 6, 4, 17, 2),
        Task(8, 7, 2, 15, 3),
        Task(9, 8, 3, 18, 1),
        Task(10, 9, 1, 9, 2),
        Task(11, 10, 4, 19, 3),
        Task(12, 11, 2, 20, 1),
        Task(13, 12, 3, 21, 4),
        Task(14, 13, 1, 8, 2),
        Task(15, 14, 4, 22, 3),
        Task(16, 15, 2, 23, 1),
        Task(17, 16, 3, 24, 4),
        Task(18, 17, 1, 7, 2),
        Task(19, 18, 2, 25, 3)
    ],

    # Task set 2 with size = 19
    [
        Task(1, 0, 3, 14, 4),
        Task(2, 1, 2, 12, 1),
        Task(3, 2, 4, 17, 3),
        Task(4, 3, 1, 10, 2),
        Task(5, 4, 3, 15, 4),
        Task(6, 5, 2, 13, 1),
        Task(7, 6, 3, 18, 3),
        Task(8, 7, 1, 9, 2),
        Task(9, 8, 4, 19, 4),
        Task(10, 9, 2, 11, 1),
        Task(11, 10, 3, 20, 2),
        Task(12, 11, 1, 8, 3),
        Task(13, 12, 4, 21, 1),
        Task(14, 13, 2, 22, 4),
        Task(15, 14, 3, 23, 2),
        Task(16, 15, 1, 7, 3),
        Task(17, 16, 4, 24, 1),
        Task(18, 17, 2, 25, 4),
        Task(19, 18, 3, 26, 2)
    ],
    # Task set 3 with size = 19
    [
        Task(1, 0, 3, 14, 4),
        Task(2, 1, 2, 12, 1),
        Task(3, 2, 4, 17, 1),
        Task(4, 3, 1, 10, 2),
        Task(5, 4, 3, 15, 4),
        Task(6, 5, 8, 13, 2),
        Task(7, 6, 3, 18, 3),
        Task(8, 7, 1, 9, 2),
        Task(9, 8, 5, 19, 4),
        Task(10, 9, 2, 11, 1),
        Task(11, 10, 7, 20, 2),
        Task(12, 11, 3, 8, 2),
        Task(13, 12, 4, 17, 1),
        Task(14, 13, 13, 20, 4),
        Task(15, 14, 3, 23, 2),
        Task(16, 15, 3, 7, 3),
        Task(17, 16, 10, 14, 1),
        Task(18, 17, 14, 25, 4),
        Task(19, 18, 3, 26, 2)
    ],

    # Task set 1 with size = 20
    [
        Task(1, 0, 2, 10, 2),
        Task(2, 1, 3, 15, 3),
        Task(3, 2, 4, 12, 1),
        Task(4, 3, 1, 8, 4),
        Task(5, 4, 2, 9, 2),
        Task(6, 5, 3, 13, 1),
        Task(7, 6, 4, 10, 3),
        Task(8, 7, 1, 9, 4),
        Task(9, 8, 2, 8, 2),
        Task(10, 9, 3, 9, 1),
        Task(11, 10, 4, 10, 3),
        Task(12, 11, 1, 8, 4),
        Task(13, 12, 3, 11, 2),
        Task(14, 13, 2, 13, 1),
        Task(15, 14, 4, 17, 3),
        Task(16, 15, 1, 9, 2),
        Task(17, 16, 3, 14, 4),
        Task(18, 17, 2, 23, 1),
        Task(19, 18, 4, 12, 2),
        Task(20, 19, 3, 14, 4)
    ],

    # Task set 2 with size = 20
    [
        Task(1, 0, 1, 9, 4),
        Task(2, 1, 3, 15, 2),
        Task(3, 2, 2, 13, 3),
        Task(4, 3, 4, 12, 1),
        Task(5, 4, 3, 16, 2),
        Task(6, 5, 1, 10, 3),
        Task(7, 6, 2, 14, 4),
        Task(8, 7, 3, 18, 1),
        Task(9, 8, 4, 11, 2),
        Task(10, 9, 1, 8, 3),
        Task(11, 10, 2, 9, 4),
        Task(12, 11, 3, 16, 1),
        Task(13, 12, 4, 12, 2),
        Task(14, 13, 1, 7, 3),
        Task(15, 14, 2, 23, 4),
        Task(16, 15, 3, 14, 1),
        Task(17, 16, 4, 12, 2),
        Task(18, 17, 1, 6, 3),
        Task(19, 18, 2, 10, 4),
        Task(20, 19, 3, 17, 1)
    ],
    # Task set 3 with size = 20
    [
        Task(1, 0, 1, 9, 4),
        Task(2, 1, 3, 15, 2),
        Task(3, 2, 2, 6, 3),
        Task(4, 13, 4, 17, 1),
        Task(5, 4, 3, 16, 2),
        Task(6, 5, 1, 10, 3),
        Task(7, 6, 2, 14, 4),
        Task(8, 7, 3, 18, 1),
        Task(9, 8, 4, 12, 2),
        Task(10, 9, 1, 10, 3),
        Task(11, 10, 2, 11, 4),
        Task(12, 11, 3, 12, 1),
        Task(13, 12, 4, 16, 2),
        Task(14, 13, 1, 7, 3),
        Task(15, 14, 2, 7, 4),
        Task(16, 15, 3, 14, 1),
        Task(17, 16, 5, 19, 2),
        Task(18, 17, 1, 6, 3),
        Task(19, 18, 2, 8, 4),
        Task(20, 19, 3, 19, 1)
    ],

]

import pandas as pd

def calculate_aperiodic_tightness(task_set: list[Task]) -> float:
    total_utilization = sum(task.burst_time / task.deadline for task in task_set)
    return round(total_utilization, 4)





def log_results(policy, taskset_size, preemptions, deadline_misses, data_transfer_count, utilization=0):
    filename = "../simulation_results_new.xlsx"
    headers = ["Policy", "Taskset Size", "Preemptions", "Deadline Misses", "Data Transfers", "Utilization"]

    # Load or create workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    # Append new row
    ws.append([policy, taskset_size, preemptions, deadline_misses, data_transfer_count, round(utilization, 4)])
    wb.save(filename)



# tasksZ = [
#     # Task Set 1: 4 tasks, Schedulable
#     [
#         Task(1, 0, 1, 10, 1),
#         Task(2, 1, 2, 9, 2),
#         Task(3, 2, 2, 8, 3),
#         Task(4, 3, 1, 7, 4),
#     ],
#
#     # Task Set 2: 5 tasks, Schedulable
#     [
#         Task(1, 0, 1, 15, 1),
#         Task(2, 1, 2, 14, 2),
#         Task(3, 2, 3, 13, 3),
#         Task(4, 3, 1, 12, 4),
#         Task(5, 4, 1, 11, 5),
#     ],
#
#     # Task Set 3: 6 tasks, Unschedulable
#     [
#         Task(1, 0, 5, 10, 1),
#         Task(2, 1, 4, 9, 2),
#         Task(3, 2, 3, 8, 3),
#         Task(4, 3, 2, 7, 4),
#         Task(5, 4, 2, 6, 5),
#         Task(6, 5, 2, 5, 6),
#     ],
#
#     # Task Set 4: 7 tasks, Schedulable
#     [
#         Task(1, 0, 1, 20, 1),
#         Task(2, 1, 2, 18, 2),
#         Task(3, 2, 3, 16, 3),
#         Task(4, 3, 4, 14, 4),
#         Task(5, 4, 1, 12, 5),
#         Task(6, 5, 2, 10, 6),
#         Task(7, 6, 3, 8, 7),
#     ],
#
#     # Task Set 5: 8 tasks, Unschedulable
#     [
#         Task(1, 0, 5, 10, 1),
#         Task(2, 1, 5, 11, 2),
#         Task(3, 2, 4, 12, 3),
#         Task(4, 3, 4, 13, 4),
#         Task(5, 4, 4, 14, 5),
#         Task(6, 5, 3, 15, 6),
#         Task(7, 6, 3, 16, 7),
#         Task(8, 7, 3, 17, 8),
#     ],
#
#     # Task Set 6: 9 tasks, Schedulable
#     [
#         Task(1, 0, 1, 20, 1),
#         Task(2, 1, 1, 19, 2),
#         Task(3, 2, 1, 18, 3),
#         Task(4, 3, 2, 17, 4),
#         Task(5, 4, 2, 16, 5),
#         Task(6, 5, 3, 15, 6),
#         Task(7, 6, 3, 14, 7),
#         Task(8, 7, 4, 13, 8),
#         Task(9, 8, 4, 12, 9),
#     ],
#
#     # Task Set 7: 10 tasks, Schedulable
#     [
#         Task(1, 0, 1, 30, 1),
#         Task(2, 1, 1, 29, 2),
#         Task(3, 2, 1, 28, 3),
#         Task(4, 3, 2, 27, 4),
#         Task(5, 4, 2, 26, 5),
#         Task(6, 5, 2, 25, 6),
#         Task(7, 6, 3, 24, 7),
#         Task(8, 7, 3, 23, 8),
#         Task(9, 8, 4, 22, 9),
#         Task(10, 9, 4, 21, 10),
#     ],
#
#     # Task Set 8: 11 tasks, Unschedulable
#     [
#         Task(1, 0, 5, 12, 1),
#         Task(2, 1, 5, 11, 2),
#         Task(3, 2, 5, 10, 3),
#         Task(4, 3, 4, 9, 4),
#         Task(5, 4, 4, 8, 5),
#         Task(6, 5, 4, 7, 6),
#         Task(7, 6, 3, 6, 7),
#         Task(8, 7, 3, 5, 8),
#         Task(9, 8, 3, 4, 9),
#         Task(10, 9, 3, 3, 10),
#         Task(11, 10, 3, 2, 11),
#     ],
#
#     # Task Set 9: 12 tasks, Schedulable
#     [
#         Task(1, 0, 2, 30, 1),
#         Task(2, 2, 3, 28, 2),
#         Task(3, 4, 2, 26, 3),
#         Task(4, 6, 3, 24, 4),
#         Task(5, 8, 4, 22, 5),
#         Task(6, 10, 2, 20, 6),
#         Task(7, 12, 1, 18, 7),
#         Task(8, 14, 1, 16, 8),
#         Task(9, 16, 1, 14, 9),
#         Task(10, 18, 1, 12, 10),
#         Task(11, 20, 1, 10, 11),
#         Task(12, 22, 1, 8, 12),
#     ],
#
#     # Task Set 10: 13 tasks, Unschedulable
#     [
#         Task(1, 0, 5, 10, 1),
#         Task(2, 1, 5, 11, 2),
#         Task(3, 2, 5, 12, 3),
#         Task(4, 3, 4, 13, 4),
#         Task(5, 4, 4, 14, 5),
#         Task(6, 5, 4, 15, 6),
#         Task(7, 6, 3, 16, 7),
#         Task(8, 7, 3, 17, 8),
#         Task(9, 8, 3, 18, 9),
#         Task(10, 9, 3, 19, 10),
#         Task(11, 10, 3, 20, 11),
#         Task(12, 11, 3, 21, 12),
#         Task(13, 12, 3, 22, 13),
#     ],
#     [
#         Task(1, 0, 1, 30, 1),
#         Task(2, 1, 1, 28, 2),
#         Task(3, 2, 1, 26, 3),
#         Task(4, 3, 2, 24, 4),
#         Task(5, 4, 2, 22, 5),
#         Task(6, 5, 3, 20, 6),
#         Task(7, 6, 3, 18, 7),
#         Task(8, 7, 4, 16, 8),
#         Task(9, 8, 4, 14, 9),
#         Task(10, 9, 4, 12, 10),
#         Task(11, 10, 2, 10, 11),
#         Task(12, 11, 1, 8, 12),
#         Task(13, 12, 1, 6, 13),
#         Task(14, 13, 1, 4, 14),
#     ],
#
#     # Task Set 12: 15 tasks, Unschedulable
#     [
#         Task(1, 0, 6, 12, 1),
#         Task(2, 1, 6, 11, 2),
#         Task(3, 2, 6, 10, 3),
#         Task(4, 3, 5, 9, 4),
#         Task(5, 4, 5, 8, 5),
#         Task(6, 5, 5, 7, 6),
#         Task(7, 6, 4, 6, 7),
#         Task(8, 7, 4, 5, 8),
#         Task(9, 8, 4, 4, 9),
#         Task(10, 9, 4, 3, 10),
#         Task(11, 10, 4, 2, 11),
#         Task(12, 11, 3, 1, 12),
#         Task(13, 12, 2, 1, 13),
#         Task(14, 13, 2, 1, 14),
#         Task(15, 14, 2, 1, 15),
#     ],
#
#     # Task Set 13: 16 tasks, Schedulable
#     [
#         Task(1, 0, 2, 40, 1),
#         Task(2, 2, 2, 38, 2),
#         Task(3, 4, 2, 36, 3),
#         Task(4, 6, 2, 34, 4),
#         Task(5, 8, 2, 32, 5),
#         Task(6, 10, 2, 30, 6),
#         Task(7, 12, 2, 28, 7),
#         Task(8, 14, 2, 26, 8),
#         Task(9, 16, 2, 24, 9),
#         Task(10, 18, 2, 22, 10),
#         Task(11, 20, 2, 20, 11),
#         Task(12, 22, 2, 18, 12),
#         Task(13, 24, 2, 16, 13),
#         Task(14, 26, 2, 14, 14),
#         Task(15, 28, 2, 12, 15),
#         Task(16, 30, 2, 10, 16),
#     ],
#
#     # Task Set 14: 17 tasks, Unschedulable
#     [
#         Task(1, 0, 10, 10, 1),
#         Task(2, 1, 10, 11, 2),
#         Task(3, 2, 10, 12, 3),
#         Task(4, 3, 10, 13, 4),
#         Task(5, 4, 10, 14, 5),
#         Task(6, 5, 9, 15, 6),
#         Task(7, 6, 8, 16, 7),
#         Task(8, 7, 8, 17, 8),
#         Task(9, 8, 7, 18, 9),
#         Task(10, 9, 6, 19, 10),
#         Task(11, 10, 6, 20, 11),
#         Task(12, 11, 5, 21, 12),
#         Task(13, 12, 5, 22, 13),
#         Task(14, 13, 5, 23, 14),
#         Task(15, 14, 5, 24, 15),
#         Task(16, 15, 5, 25, 16),
#         Task(17, 16, 5, 26, 17),
#     ],
#
#     # Task Set 15: 18 tasks, Schedulable
#     [
#         Task(1, 0, 1, 60, 1),
#         Task(2, 1, 1, 58, 2),
#         Task(3, 2, 1, 56, 3),
#         Task(4, 3, 2, 54, 4),
#         Task(5, 4, 2, 52, 5),
#         Task(6, 5, 3, 50, 6),
#         Task(7, 6, 3, 48, 7),
#         Task(8, 7, 4, 46, 8),
#         Task(9, 8, 4, 44, 9),
#         Task(10, 9, 5, 42, 10),
#         Task(11, 10, 5, 40, 11),
#         Task(12, 11, 5, 38, 12),
#         Task(13, 12, 6, 36, 13),
#         Task(14, 13, 6, 34, 14),
#         Task(15, 14, 6, 32, 15),
#         Task(16, 15, 7, 30, 16),
#         Task(17, 16, 7, 28, 17),
#         Task(18, 17, 7, 26, 18),
#     ],
#
#     # Task Set 16: 19 tasks, Unschedulable
#     [
#         Task(1, 0, 15, 30, 1),
#         Task(2, 1, 15, 30, 2),
#         Task(3, 2, 15, 30, 3),
#         Task(4, 3, 15, 30, 4),
#         Task(5, 4, 15, 30, 5),
#         Task(6, 5, 15, 30, 6),
#         Task(7, 6, 15, 30, 7),
#         Task(8, 7, 15, 30, 8),
#         Task(9, 8, 15, 30, 9),
#         Task(10, 9, 15, 30, 10),
#         Task(11, 10, 15, 30, 11),
#         Task(12, 11, 15, 30, 12),
#         Task(13, 12, 15, 30, 13),
#         Task(14, 13, 15, 30, 14),
#         Task(15, 14, 15, 30, 15),
#         Task(16, 15, 15, 30, 16),
#         Task(17, 16, 15, 30, 17),
#         Task(18, 17, 15, 30, 18),
#         Task(19, 18, 15, 30, 19),
#     ],
#     [
#         Task(1, 0, 2, 80, 1),
#         Task(2, 2, 2, 78, 2),
#         Task(3, 4, 2, 76, 3),
#         Task(4, 6, 2, 74, 4),
#         Task(5, 8, 2, 72, 5),
#         Task(6, 10, 2, 70, 6),
#         Task(7, 12, 2, 68, 7),
#         Task(8, 14, 2, 66, 8),
#         Task(9, 16, 2, 64, 9),
#         Task(10, 18, 2, 62, 10),
#         Task(11, 20, 2, 60, 11),
#         Task(12, 22, 2, 58, 12),
#         Task(13, 24, 2, 56, 13),
#         Task(14, 26, 2, 54, 14),
#         Task(15, 28, 2, 52, 15),
#         Task(16, 30, 2, 50, 16),
#         Task(17, 32, 2, 48, 17),
#         Task(18, 34, 2, 46, 18),
#         Task(19, 36, 2, 44, 19),
#         Task(20, 38, 2, 42, 20),
#     ],
#
#     # Task Set 18: 21 tasks, Unschedulable
#     [
#         Task(1, 0, 10, 30, 1),
#         Task(2, 1, 10, 29, 2),
#         Task(3, 2, 10, 28, 3),
#         Task(4, 3, 10, 27, 4),
#         Task(5, 4, 10, 26, 5),
#         Task(6, 5, 10, 25, 6),
#         Task(7, 6, 10, 24, 7),
#         Task(8, 7, 10, 23, 8),
#         Task(9, 8, 10, 22, 9),
#         Task(10, 9, 10, 21, 10),
#         Task(11, 10, 10, 20, 11),
#         Task(12, 11, 10, 19, 12),
#         Task(13, 12, 10, 18, 13),
#         Task(14, 13, 10, 17, 14),
#         Task(15, 14, 10, 16, 15),
#         Task(16, 15, 10, 15, 16),
#         Task(17, 16, 10, 14, 17),
#         Task(18, 17, 10, 13, 18),
#         Task(19, 18, 10, 12, 19),
#         Task(20, 19, 10, 11, 20),
#         Task(21, 20, 10, 10, 21),
#     ],
#
#     # Task Set 19: 22 tasks, Schedulable
#     [
#         Task(1, 0, 1, 100, 1),
#         Task(2, 1, 1, 98, 2),
#         Task(3, 2, 1, 96, 3),
#         Task(4, 3, 1, 94, 4),
#         Task(5, 4, 2, 92, 5),
#         Task(6, 5, 2, 90, 6),
#         Task(7, 6, 2, 88, 7),
#         Task(8, 7, 2, 86, 8),
#         Task(9, 8, 2, 84, 9),
#         Task(10, 9, 3, 82, 10),
#         Task(11, 10, 3, 80, 11),
#         Task(12, 11, 3, 78, 12),
#         Task(13, 12, 3, 76, 13),
#         Task(14, 13, 4, 74, 14),
#         Task(15, 14, 4, 72, 15),
#         Task(16, 15, 4, 70, 16),
#         Task(17, 16, 4, 68, 17),
#         Task(18, 17, 5, 66, 18),
#         Task(19, 18, 5, 64, 19),
#         Task(20, 19, 5, 62, 20),
#         Task(21, 20, 5, 60, 21),
#         Task(22, 21, 5, 58, 22),
#     ],
#
#     # Task Set 20: 23 tasks, Unschedulable
#     [
#         Task(1, 0, 10, 20, 1),
#         Task(2, 1, 10, 21, 2),
#         Task(3, 2, 10, 22, 3),
#         Task(4, 3, 10, 23, 4),
#         Task(5, 4, 10, 24, 5),
#         Task(6, 5, 10, 25, 6),
#         Task(7, 6, 10, 26, 7),
#         Task(8, 7, 10, 27, 8),
#         Task(9, 8, 10, 28, 9),
#         Task(10, 9, 10, 29, 10),
#         Task(11, 10, 10, 30, 11),
#         Task(12, 11, 10, 31, 12),
#         Task(13, 12, 10, 32, 13),
#         Task(14, 13, 10, 33, 14),
#         Task(15, 14, 10, 34, 15),
#         Task(16, 15, 10, 35, 16),
#         Task(17, 16, 10, 36, 17),
#         Task(18, 17, 10, 37, 18),
#         Task(19, 18, 10, 38, 19),
#         Task(20, 19, 10, 39, 20),
#         Task(21, 20, 10, 40, 21),
#         Task(22, 21, 10, 41, 22),
#         Task(23, 22, 10, 42, 23),
#     ],
# ]
# tasks2 = [
#     # Task Set 1: 4 tasks, Schedulable
#     [
#         Task(1, 0, 1, 10, 1),
#         Task(2, 1, 2, 9, 2),
#         Task(3, 2, 2, 8, 3),
#         Task(4, 3, 1, 7, 4),
#     ],
#     # Task Set 2: 4 tasks, Non-schedulable
#     [
#         Task(1, 0, 3, 5, 1),
#         Task(2, 1, 3, 6, 2),
#         Task(3, 2, 3, 7, 3),
#         Task(4, 3, 3, 8, 4),
#     ],
#     # Task Set 3: 4 tasks, Schedulable
#     [
#         Task(1, 0, 1, 12, 1),
#         Task(2, 1, 1, 11, 2),
#         Task(3, 2, 1, 10, 3),
#         Task(4, 3, 1, 9, 4),
#     ],
#     # Task Set 4: 5 tasks, Schedulable
#     [
#         Task(1, 0, 1, 15, 1),
#         Task(2, 1, 2, 14, 2),
#         Task(3, 2, 3, 13, 3),
#         Task(4, 3, 1, 12, 4),
#         Task(5, 4, 1, 11, 5),
#     ],
#     # Task Set 5: 5 tasks, Non-schedulable
#     [
#         Task(1, 0, 4, 8, 1),
#         Task(2, 1, 4, 9, 2),
#         Task(3, 2, 4, 10, 3),
#         Task(4, 3, 4, 11, 4),
#         Task(5, 4, 4, 12, 5),
#     ],
#     # Task Set 6: 5 tasks, Schedulable
#     [
#         Task(1, 0, 1, 16, 1),
#         Task(2, 1, 1, 15, 2),
#         Task(3, 2, 1, 14, 3),
#         Task(4, 3, 1, 13, 4),
#         Task(5, 4, 1, 12, 5),
#     ],
#     # Task Set 7: 6 tasks, Non-schedulable
#     [
#         Task(1, 0, 5, 10, 1),
#         Task(2, 1, 4, 9, 2),
#         Task(3, 2, 3, 8, 3),
#         Task(4, 3, 2, 7, 4),
#         Task(5, 4, 2, 6, 5),
#         Task(6, 5, 2, 5, 6),
#     ],
#     # Task Set 8: 6 tasks, Schedulable
#     [
#         Task(1, 0, 1, 20, 1),
#         Task(2, 1, 2, 18, 2),
#         Task(3, 2, 3, 16, 3),
#         Task(4, 3, 4, 14, 4),
#         Task(5, 4, 1, 12, 5),
#         Task(6, 5, 2, 10, 6),
#     ],
#     # Task Set 9: 6 tasks, Non-schedulable
#     [
#         Task(1, 0, 4, 8, 1),
#         Task(2, 1, 4, 9, 2),
#         Task(3, 2, 4, 10, 3),
#         Task(4, 3, 4, 11, 4),
#         Task(5, 4, 4, 12, 5),
#         Task(6, 5, 4, 13, 6),
#     ],
#     # Task Set 10: 7 tasks, Schedulable
#     [
#         Task(1, 0, 1, 30, 1),
#         Task(2, 1, 1, 29, 2),
#         Task(3, 2, 1, 28, 3),
#         Task(4, 3, 2, 27, 4),
#         Task(5, 4, 2, 26, 5),
#         Task(6, 5, 2, 25, 6),
#         Task(7, 6, 3, 24, 7),
#     ],
#     # Task Set 11: 7 tasks, Non-schedulable
#     [
#         Task(1, 0, 5, 12, 1),
#         Task(2, 1, 5, 11, 2),
#         Task(3, 2, 5, 10, 3),
#         Task(4, 3, 4, 9, 4),
#         Task(5, 4, 4, 8, 5),
#         Task(6, 5, 4, 7, 6),
#         Task(7, 6, 3, 6, 7),
#     ],
#     # Task Set 12: 7 tasks, Schedulable
#     [
#         Task(1, 0, 1, 25, 1),
#         Task(2, 1, 1, 24, 2),
#         Task(3, 2, 1, 23, 3),
#         Task(4, 3, 2, 22, 4),
#         Task(5, 4, 2, 21, 5),
#         Task(6, 5, 2, 20, 6),
#         Task(7, 6, 3, 19, 7),
#     ],
#     # Task Set 13: 8 tasks, Non-schedulable
#     [
#         Task(1, 0, 5, 10, 1),
#         Task(2, 1, 5, 11, 2),
#         Task(3, 2, 4, 12, 3),
#         Task(4, 3, 4, 13, 4),
#         Task(5, 4, 4, 14, 5),
#         Task(6, 5, 3, 15, 6),
#         Task(7, 6, 3, 16, 7),
#         Task(8, 7, 3, 17, 8),
#     ],
#     # Task Set 14: 8 tasks, Schedulable
#     [
#         Task(1, 0, 1, 20, 1),
#         Task(2, 1, 1, 19, 2),
#         Task(3, 2, 1, 18, 3),
#         Task(4, 3, 2, 17, 4),
#         Task(5, 4, 2, 16, 5),
#         Task(6, 5, 3, 15, 6),
#         Task(7, 6, 3, 14, 7),
#         Task(8, 7, 4, 13, 8),
#     ],
#     # Task Set 15: 8 tasks, Non-schedulable
#     [
#         Task(1, 0, 4, 8, 1),
#         Task(2, 1, 4, 9, 2),
#         Task(3, 2, 4, 10, 3),
#         Task(4, 3, 4, 11, 4),
#         Task(5, 4, 4, 12, 5),
#         Task(6, 5, 4, 13, 6),
#         Task(7, 6, 4, 14, 7),
#         Task(8, 7, 4, 15, 8),
#     ],
#     # Task Set 16: 9 tasks, Schedulable
#     [
#         Task(1, 0, 1, 20, 1),
#         Task(2, 1, 1, 19, 2),
#         Task(3, 2, 1, 18, 3),
#         Task(4, 3, 2, 17, 4),
#         Task(5, 4, 2, 16, 5),
#         Task(6, 5, 3, 15, 6),
#         Task(7, 6, 3, 14, 7),
#         Task(8, 7, 4, 13, 8),
#         Task(9, 8, 4, 12, 9),
#     ],
#     # Task Set 17: 9 tasks, Non-schedulable
#     [
#         Task(1, 0, 5, 10, 1),
#         Task(2, 1, 5, 11, 2),
#         Task(3, 2, 5, 12, 3),
#         Task(4, 3, 4, 13, 4),
#         Task(5, 4, 4, 14, 5),
#         Task(6, 5, 4, 15, 6),
#         Task(7, 6, 3, 16, 7),
#         Task(8, 7, 3, 17, 8),
#         Task(9, 8, 3, 18, 9),
#     ],
#     # Task Set 18: 9 tasks, Schedulable
#     [
#         Task(1, 0, 1, 30, 1),
#         Task(2, 1, 1, 29, 2),
#         Task(3, 2, 1, 28, 3),
#         Task(4, 3, 2, 27, 4),
#         Task(5, 4, 2, 26, 5),
#         Task(6, 5, 2, 25, 6),
#         Task(7, 6, 3, 24, 7),
#         Task(8, 7, 3, 23, 8),
#         Task(9, 8, 4, 22, 9),
#     ],
#     # Task Set 19: 10 tasks, Non-schedulable
#     [
#         Task(1, 0, 10, 10, 1),
#         Task(2, 1, 10, 11, 2),
#         Task(3, 2, 10, 12, 3),
#         Task(4, 3, 10, 13, 4),
#         Task(5, 4, 10, 14, 5),
#         Task(6, 5, 9, 15, 6),
#         Task(7, 6, 8, 16, 7),
#         Task(8, 7, 8, 17, 8),
#         Task(9, 8, 7, 18, 9),
#         Task(10, 9, 6, 19, 10),
#     ],
#     # Task Set 20: 10 tasks, Schedulable
#     [
#         Task(1, 0, 1, 60, 1),
#         Task(2, 1, 1, 58, 2),
#         Task(3, 2, 1, 56, 3),
#         Task(4, 3, 2, 54, 4),
#         Task(5, 4, 2, 52, 5),
#         Task(6, 5, 3, 50, 6),
#         Task(7, 6, 3, 48, 7),
#         Task(8, 7, 4, 46, 8),
#         Task(9, 8, 4, 44, 9),
#         Task(10, 9, 5, 42, 10),
#     ],
#     # Task Set 21: 10 tasks, Non-schedulable
#     [
#         Task(1, 0, 4, 8, 1),
#         Task(2, 1, 4, 9, 2),
#         Task(3, 2, 4, 10, 3),
#         Task(4, 3, 4, 11, 4),
#         Task(5, 4, 4, 12, 5),
#         Task(6, 5, 4, 13, 6),
#         Task(7, 6, 4, 14, 7),
#         Task(8, 7, 4, 15, 8),
#         Task(9, 8, 4, 16, 9),
#         Task(10, 9, 4, 17, 10),
#     ],
# ]
# tasksZ = [
#     # Task Set 1 (3 tasks)
#     [
#         Task("T1", 0, 2, 5, 3),
#         Task("T2", 1, 3, 8, 2),
#         Task("T3", 2, 2, 7, 1),
#     ],
#     # Task Set 2 Task(4 tasks)
#     [
#         Task("T1", 0, 4, 10, 2),
#         Task("T2", 1, 2, 7, 3),
#         Task("T3", 3, 3, 9, 1),
#         Task("T4", 5, 1, 6, 2),
#     ],
#     # Task Set 3 Task(5 tasks)
#     [
#         Task("T1", 0, 3, 8, 1),
#         Task("T2", 2, 2, 6, 2),
#         Task("T3", 4, 1, 7, 3),
#         Task("T4", 5, 3, 10, 2),
#         Task("T5", 6, 2, 9, 1),
#     ],
#     # Task Set 4 Task(6 tasks)
#     [
#         Task("T1", 1, 3, 8, 1),
#         Task("T2", 2, 2, 7, 2),
#         Task("T3", 4, 3, 9, 3),
#         Task("T4", 5, 2, 8, 2),
#         Task("T5", 6, 1, 7, 1),
#         Task("T6", 7, 4, 12, 2),
#     ],
#     # Task Set 5 Task(7 tasks)
#     [
#         Task("T1", 0, 2, 6, 1),
#         Task("T2", 1, 3, 9, 2),
#         Task("T3", 2, 2, 7, 3),
#         Task("T4", 4, 1, 6, 2),
#         Task("T5", 5, 3, 10, 1),
#         Task("T6", 6, 2, 8, 3),
#         Task("T7", 7, 4, 13, 2),
#     ],
#     # Task Set 6 Task(8 tasks)
#     [
#         Task("T1", 0, 2, 6, 2),
#         Task("T2", 2, 3, 10, 1),
#         Task("T3", 3, 2, 8, 3),
#         Task("T4", 4, 4, 12, 2),
#         Task("T5", 6, 1, 7, 1),
#         Task("T6", 7, 3, 11, 3),
#         Task("T7", 8, 2, 10, 2),
#         Task("T8", 9, 4, 14, 1),
#     ],
#     # Task Set 7 Task(9 tasks)
#     [
#         Task("T1", 1, 3, 8, 1),
#         Task("T2", 2, 4, 11, 2),
#         Task("T3", 3, 2, 9, 3),
#         Task("T4", 4, 1, 7, 1),
#         Task("T5", 6, 3, 10, 2),
#         Task("T6", 7, 2, 9, 3),
#         Task("T7", 8, 4, 14, 1),
#         Task("T8", 9, 2, 11, 2),
#         Task("T9", 10, 1, 8, 3),
#     ],
#     # Task Set 8 Task(10 tasks)
#     [
#         Task("T1", 0, 2, 6, 1),
#         Task("T2", 1, 3, 9, 2),
#         Task("T3", 2, 2, 7, 3),
#         Task("T4", 3, 1, 6, 1),
#         Task("T5", 5, 4, 12, 2),
#         Task("T6", 6, 3, 10, 3),
#         Task("T7", 7, 2, 9, 2),
#         Task("T8", 8, 3, 11, 1),
#         Task("T9", 9, 4, 14, 2),
#         Task("T10", 10, 2, 12, 3),
#     ],
#     # Task Set 9 Task(3 tasks)
#     [
#         Task("T1", 0, 3, 8, 1),
#         Task("T2", 1, 2, 6, 2),
#         Task("T3", 2, 1, 5, 3),
#     ],
#     # Task Set 10 Task(5 tasks)
#     [
#         Task("T1", 0, 4, 10, 1),
#         Task("T2", 1, 3, 8, 2),
#         Task("T3", 3, 2, 7, 3),
#         Task("T4", 5, 1, 6, 1),
#         Task("T5", 6, 3, 9, 2),
#     ],
# ]


def main():

    all_tasks = [task for sublist in tasks for task in sublist]
    execution_times = [task.burst_time for task in all_tasks]
    deadlines = [task.deadline for task in all_tasks]
    # Calculations
    number_of_tasks = len(all_tasks)
    average_deadline = np.mean(deadlines)
    # deadline_tightness = (np.mean([et / dl for et, dl in zip(execution_times, deadlines)])) * 100  # Convert to percentage

    # Displaying results
    print("Metric Table:")
    print(f"Number of Tasks: {number_of_tasks}")
    print(f"Average Deadline: {average_deadline:.2f} ms")
    # print(f"Deadline Tightness: {deadline_tightness:.2f}%")

    for i, taskset_ in enumerate(tasks):
        for t in taskset_:
            if not hasattr(t, 'utilization'):
                t.utilization = t.burst_time / t.deadline if t.deadline > 0 else 0.0

    # Calculate tightness for all sets
    aperiodic_tightness_results = []
    for idx, task_set in enumerate(tasks):
        tightness = calculate_aperiodic_tightness(task_set)
        aperiodic_tightness_results.append(
            # "Task Set": idx+1,
            # "Number of Tasks": len(task_set),
            tightness)
    # Create table
    aperiodic_tightness_df = pd.DataFrame(aperiodic_tightness_results)
    print(aperiodic_tightness_df)
    print("tightness:", np.mean(aperiodic_tightness_results))


if __name__ == "__main__":
    main()
